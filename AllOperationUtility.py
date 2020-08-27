#!/usr/bin/env python
# coding: utf-8

# In[1]:


from sendingEmails import EmailSendingUtility
import yaml
import openpyxl as xl;
import math
import xlsxwriter
import logging
import smtplib
from FileDetails import FileDetailsUtility 
from sendingEmails import EmailSendingUtility
from smtplib import * 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 
import os, uuid, sys
from azure.storage.blob import BlobServiceClient, PublicAccess
import pysftp


# In[2]:


class SplitExcelOperation:
    def splitExcel(self,total_splits,sourcefile,splitRatioArray,sumOfSplits,senderIdPwd,vendorId,asympCode,sympCode):
        self.start_row_source = []
        excelList =[]
        logging.debug("Inside Split Excel")
        
        sourcefile = self.insertCampainCode(sourcefile,asympCode,sympCode)
        
        for sheetcount in range(0,len(sourcefile.sheetnames)):
            self.start_row_source=2
            #print(sourcefile.sheetnames)
            logging.debug("Sheet names of source file" , sourcefile.sheetnames)
            sourceFile = sourcefile.worksheets[sheetcount]
            
            for j in range(0,total_splits):
                self.newWorkBook = xl.Workbook()
                self.split_row_list=[]   
                self.newWorkBookSheet =self.newWorkBook.active
                
                maxRowAndCol=[sourceFile.max_row,sourceFile.max_column]

                for k in range(1,maxRowAndCol[1]+1):
                    self.firstRow_source = sourceFile.cell(row = 1, column = k) 
                    self.newWorkBookSheet.cell(row = 1, column = k).value = self.firstRow_source.value 

                self.split_ratio_row = math.ceil((splitRatioArray[j]/sumOfSplits)*(maxRowAndCol[0]-1))
                #self.split_row_list.append(self.split_ratio_row)
                self.start_row_newsheet=2
                logging.debug('Rows Split: ',self.split_ratio_row)

                for k in range (self.start_row_source, self.start_row_source+self.split_ratio_row): 
                    for l in range(1,maxRowAndCol[1]+1):
                        self.firstRow_source = sourceFile.cell(row = k, column = l) 
                        self.newWorkBookSheet.cell(row =self.start_row_newsheet, column = l).value= self.firstRow_source.value

                    self.start_row_newsheet =self.start_row_newsheet + 1

                self.start_row_source = self.start_row_source + self.split_ratio_row
                logging.debug('Start source',self.start_row_source)
                self.fileNameNewWorkbook="python_files/report_file"+str(uuid.uuid4())+".xlsx"
                self.newWorkBook.save(self.fileNameNewWorkbook)
                excelList.append(self.fileNameNewWorkbook)
            #self.storeAsBlob()
        self.sendEmail(excelList,senderIdPwd,vendorId)
        return excelList
        
    def insertCampainCode(self,sourceFiles,asympCode,sympCode):
         
        for sheetcount in range(0,len(sourceFiles.sheetnames)):
            sourceFileSheet = sourceFiles.worksheets[sheetcount]
            max_column = sourceFileSheet.max_column+1
            sourceFileSheet.cell(row = 1,column = max_column).value = "Campaign Code"
           
            if(sourceFileSheet.cell(row = 2,column = 1).value.lower()=="asymptomatic"):
                for i in range (2,sourceFileSheet.max_row+1):
                    sourceFileSheet.cell(row = i, column = max_column).value = asympCode
            else:
                for i in range (2,sourceFileSheet.max_row+1):
                    sourceFileSheet.cell(row = i, column = max_column).value = sympCode
        return sourceFiles
            
                
    def sendEmail(self,excelList,senderIdPwd,vendorId):
        k=0
        for i in range(0,len(excelList)):
            emailUtility = EmailSendingUtility()
            if(k==len(vendorId)):
                k=0
            msgText = emailUtility.getEmailBody(senderIdPwd,vendorId,k,excelList[i])
            s = smtplib.SMTP('smtp.gmail.com', 587) 
            s.starttls() 
            s.login(str(senderIdPwd[0]),str(senderIdPwd[1])) 
            try:
                s.sendmail(str(senderIdPwd[0]), str(vendorId[k]), msgText) 
                print("mail sent successfully")
            except smtplib.SMTPException as e:
                print("Unable to send email -r" ,e)
                s.quit() 
            k=k+1
        
        

