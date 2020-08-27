#!/usr/bin/env python
# coding: utf-8

# In[1]:


import yaml
import openpyxl as xl;
import math
import xlsxwriter
import logging
import smtplib
from FileDetails import FileDetailsUtility 
from SplitExcelUtility import SplitExcel
from sendingEmails import EmailSendingUtility
from smtplib import * 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 
import os, uuid, sys
from azure.storage.blob import BlobServiceClient, PublicAccess
import pysftp
from UploadFileUtility import BlobStorageClass
from AllOperationUtility import SplitExcelOperation


# In[2]:


def main():
    fileDet = FileDetailsUtility()
    fileDet.getConfigurationalData("python_files/configuration_details.yml")
    #getConfig=GettingConfiguration("python_files/configuration_details.yml")
    totalSplits=fileDet.total_splits
    splitRatioArray=fileDet.ratio_array
    sumOfSplits=fileDet.sum_ratio_array
    senderIdPwd=fileDet.senderIdPwd
    vendorId=fileDet.vendorsId
    hostName = fileDet.hostName
    userName = fileDet.user
    password = fileDet.password
    asymptomaticCod = fileDet.asympCode;
    symptomaticCod = fileDet.sympCode
    uploadUrl = fileDet.uploadUrl

    #print('TotalSplits:',totalSplits,'Sum of Splits:',sumOfSplits,'VendorId',vendorId)
    logging.debug('TotalSplits:',totalSplits,'Sum of Splits:',sumOfSplits,'VendorId: ',vendorId)


    sourceFile = fileDet.downloadFileFromSftp(hostName,userName,password)
    #maxRowAndCol = [sourceFile.max_row,sourceFile.max_column]
    newWorkBook = fileDet.createNewWorkbook()
    #logging.debug('Max Row: ' ,maxRowAndCol[0],'Max Column: ',maxRowAndCol[1])
    #emailText = EmailSendingUtility.getEmailBody()


    #om=operationalMethods()
    splitFunc = SplitExcelOperation()
    listOfExcels=splitFunc.splitExcel(totalSplits,sourceFile,splitRatioArray,sumOfSplits,senderIdPwd,vendorId,asymptomaticCod,symptomaticCod)
    print(listOfExcels)
    blobstore=BlobStorageClass()
    blobstore.storeAsBlob(listOfExcels)
    blobstore.uploadFilePost(listOfExcels,asymptomaticCod,symptomaticCod,uploadUrl)

if __name__=="__main__": 
    main() 

