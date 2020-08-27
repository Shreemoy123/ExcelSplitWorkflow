#!/usr/bin/env python
# coding: utf-8

# In[1]:


import yaml
import pysftp
import os, uuid, sys
import openpyxl as xl;
import logging


# In[19]:


class FileDetailsUtility:
    def getConfigurationalData(arg,filename):
        with open(filename,"r") as file:
            arg.documents = yaml.full_load(file)

            for item, doc in arg.documents.items():
                #print(item, ":", doc)
                logging.debug('Item: ', item , 'value: ',doc)
            arg.total_splits=arg.documents.get('splits')
            arg.split_ratio=str(arg.documents.get('ratio'))
            arg.ratio_array=[int(i) for i in arg.split_ratio.replace(',','')]
            arg.sum_ratio_array = sum(arg.ratio_array)
            arg.senderIdPwd = arg.documents.get('senderIdPwd').split(",")
            arg.vendorsId = arg.documents.get('vendors').split(",")
            arg.hostName = arg.documents.get('host')
            arg.user = arg.documents.get('user')
            arg.password = arg.documents.get('password')
            arg.asympCode = arg.documents.get('asymptomaticCode')
            arg.sympCode = arg.documents.get('symptomaticCode')
            arg.uploadUrl = arg.documents.get('uploadUrl')
                
    def downloadFileFromSftp(self,hostName,userName,password):
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        self.fileName="python_files/SftpDownloads/DownloadedReport"+str(uuid.uuid4())+".xlsx"
        with pysftp.Connection(hostName, username=userName, password=password,cnopts=cnopts) as sftp:
            with sftp.cd('public'): 
                sftp.get('/exports/Financial_Sample.xlsx',self.fileName) 
        #return self.fileName
        self.wb1 = xl.load_workbook(self.fileName) 
        #self.source_wb=self.wb1.worksheets[1]
        return self.wb1
    
    def createNewWorkbook(self):
        self.new_wb1 = xl.Workbook() 
        return self.new_wb1 

