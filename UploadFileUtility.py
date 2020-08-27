#!/usr/bin/env python
# coding: utf-8

# In[12]:


from azure.storage.blob import BlobServiceClient, PublicAccess
import os, uuid, sys
import requests
import openpyxl as xl;


# In[16]:


class BlobStorageClass:    
    def storeAsBlob(self,excelSheets):
        connection_string ="DefaultEndpointsProtocol=https;AccountName=devinfysurvey;AccountKey=Ds9DK/roan7/l/DvW/+SjdVJhGQKuaRvPFuhK2xImkuY22WHhgMuaLKY7rAkwXlnqXImIoIv8vIfqrlBcCOAJg==;EndpointSuffix=core.windows.net"
        container_name ='dev-infy-survey-container'
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        #container_client = blob_service_client.create_container(container_name)
        for i in range(0,len(excelSheets)):
            local_path = "./data"
            self.local_file_name = os.path.basename(excelSheets[i])
            self.filename = excelSheets[i]

            blob_client = blob_service_client.get_blob_client(container=container_name, blob=self.local_file_name)

            print("\nUploading to Azure Storage as blob:\n\t" + self.local_file_name)

            # Upload the created file
            with open(self.filename, "rb") as data:
                blob_client.upload_blob(data)
                
    def uploadFilePost(self,excelSheets,asympCode,sympCode,uploadUrl):
        campaignCode = asympCode
        for i in range (0,len(excelSheets)):
            self.wb1 = xl.load_workbook(excelSheets[i])
            self.source_wb=self.wb1.worksheets[0]
            
            max_col = self.source_wb.max_column
            if(self.source_wb.cell(row = 1,column = max_col).value == asympCode):
                campaignCode=asympCode
            else:
                campaignCode = sympCode
            with open(excelSheets[i], "rb") as file:
                response = requests.post(uploadUrl,data=[{'file':'file'},{'campaignCode':'campaignCode'}])
                       
        

